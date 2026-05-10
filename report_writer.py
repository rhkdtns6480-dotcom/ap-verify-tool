"""
report_writer.py — Excel(.xlsx) + PDF 결과보고서 생성
"""
import os
import sys
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter

from scenario_engine import RunRecord, Scenario


# ──────────────────────────────────────────────
# 경로 헬퍼 (PyInstaller 대응)
# ──────────────────────────────────────────────

def _res(rel: str) -> str:
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, rel)


def _report_dir() -> str:
    if getattr(sys, "frozen", False):
        d = os.path.join(os.environ.get("APPDATA", ""), "APVerifyTool", "reports")
    else:
        d = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")
    os.makedirs(d, exist_ok=True)
    return d


# ──────────────────────────────────────────────
# 색상 상수
# ──────────────────────────────────────────────

C_HEADER   = "1F4E79"
C_PASS     = "E2EFDA"
C_FAIL     = "FCE4D6"
C_WARN     = "FFF2CC"
C_TITLE_FG = "FFFFFF"
C_BORDER   = "BFBFBF"


def _border(style="thin"):
    s = Side(border_style=style, color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color: str):
    return PatternFill("solid", fgColor=hex_color)


# ──────────────────────────────────────────────
# Excel 생성
# ──────────────────────────────────────────────

def export_excel(
    scenario: Scenario,
    records: list[RunRecord],
    device_info: dict,
    save_dir: Optional[str] = None,
) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    _write_summary_sheet(wb, scenario, records, device_info)
    _write_detail_sheet(wb, scenario, records)
    _write_log_sheet(wb, records)

    if save_dir is None:
        save_dir = _report_dir()
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    name = f"{scenario.plugin}_{scenario.name}_{ts}.xlsx"
    path = os.path.join(save_dir, name)
    wb.save(path)
    return path


def _write_summary_sheet(wb, scenario, records, device_info):
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30

    hf = Font(bold=True, color=C_TITLE_FG, size=12)
    hfi = PatternFill("solid", fgColor=C_HEADER)

    def _row(label, value, r):
        ws.cell(r, 1, label).font = Font(bold=True)
        ws.cell(r, 2, value)
        for c in (1, 2):
            ws.cell(r, c).border = _border()

    # 제목
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "AP 자동검증 결과 요약"
    c.font = Font(bold=True, size=14, color=C_TITLE_FG)
    c.fill = hfi
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    r = 3
    _row("시나리오", scenario.name, r); r+=1
    _row("플러그인", scenario.plugin, r); r+=1
    _row("장비명", device_info.get("device_name","—"), r); r+=1
    _row("펌웨어", device_info.get("firmware","—"), r); r+=1
    _row("담당자", device_info.get("tester","—"), r); r+=1
    _row("테스트 일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), r); r+=1

    r += 1
    total_pass = sum(rec.pass_count for rec in records)
    total_fail = sum(rec.fail_count for rec in records)
    total_all  = total_pass + total_fail
    rate       = f"{total_pass/total_all*100:.1f}%" if total_all else "—"

    _row("총 반복 횟수", len(records), r); r+=1
    _row("총 Step 수", total_all, r); r+=1
    ws.cell(r, 1, "PASS").font = Font(bold=True, color="375623")
    ws.cell(r, 2, total_pass).font = Font(color="375623")
    for c in (1,2): ws.cell(r,c).border = _border()
    r += 1
    ws.cell(r, 1, "FAIL").font = Font(bold=True, color="9C0006")
    ws.cell(r, 2, total_fail).font = Font(color="9C0006")
    for c in (1,2): ws.cell(r,c).border = _border()
    r += 1
    _row("합격률", rate, r)

    # 판정
    r += 2
    ws.merge_cells(f"A{r}:B{r}")
    judged = total_fail == 0
    ws[f"A{r}"].value = "최종 판정: " + ("PASS ✓" if judged else "FAIL ✗")
    ws[f"A{r}"].font  = Font(bold=True, size=13, color=("375623" if judged else "9C0006"))
    ws[f"A{r}"].fill  = PatternFill("solid", fgColor=("E2EFDA" if judged else "FCE4D6"))
    ws[f"A{r}"].alignment = Alignment(horizontal="center")


def _write_detail_sheet(wb, scenario, records):
    ws = wb.create_sheet(f"{scenario.plugin}_Result")
    headers = ["회차","Step#","유형","항목명","기댓값","실측값","허용오차","경과(ms)","판정"]
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 8

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.font  = Font(bold=True, color=C_TITLE_FG)
        cell.fill  = PatternFill("solid", fgColor=C_HEADER)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _border()

    row = 2
    for rec in records:
        for sr in rec.step_results:
            vals = [
                rec.iteration, sr.step_index+1, sr.step_type,
                sr.step_name, sr.expected or "—", sr.actual or "—",
                "—", f"{sr.elapsed_ms:.0f}", "PASS" if sr.passed else "FAIL",
            ]
            fill = _fill(C_PASS) if sr.passed else _fill(C_FAIL)
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row, ci, v)
                c.border = _border()
                c.alignment = Alignment(horizontal="center")
                if ci == 9:
                    c.fill = fill
                    c.font = Font(bold=True, color=("375623" if sr.passed else "9C0006"))
            row += 1


def _write_log_sheet(wb, records):
    ws = wb.create_sheet("RawLog")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 50
    headers = ["회차","이벤트"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.font  = Font(bold=True, color=C_TITLE_FG)
        c.fill  = PatternFill("solid", fgColor=C_HEADER)
        c.border = _border()
    row = 2
    for rec in records:
        for sr in rec.step_results:
            if not sr.passed and sr.message:
                ws.cell(row,1,f"회차 {rec.iteration}").border = _border()
                ws.cell(row,2,sr.message).border = _border()
                row += 1


# ──────────────────────────────────────────────
# PDF 생성
# ──────────────────────────────────────────────

def export_pdf(
    scenario: Scenario,
    records: list[RunRecord],
    device_info: dict,
    save_dir: Optional[str] = None,
) -> str:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                        Table, TableStyle)
    except ImportError:
        raise RuntimeError("reportlab 패키지가 필요합니다: pip install reportlab")

    # 한글 폰트 등록
    font_name = "Helvetica"
    for candidate in [
        _res("malgun.ttf"),
        r"C:\Windows\Fonts\malgun.ttf",
        r"C:\Windows\Fonts\NanumGothic.ttf",
    ]:
        if os.path.exists(candidate):
            try:
                pdfmetrics.registerFont(TTFont("Malgun", candidate))
                font_name = "Malgun"
                break
            except Exception:
                pass

    if save_dir is None:
        save_dir = _report_dir()
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    name = f"{scenario.plugin}_{scenario.name}_{ts}.pdf"
    path = os.path.join(save_dir, name)

    doc  = SimpleDocTemplate(path, pagesize=A4,
                             leftMargin=20*mm, rightMargin=20*mm,
                             topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    normal = ParagraphStyle("n", fontName=font_name, fontSize=9, leading=14)
    title_s = ParagraphStyle("t", fontName=font_name, fontSize=16, leading=22,
                              textColor=colors.HexColor(f"#1F4E79"), spaceAfter=6)
    h2_s   = ParagraphStyle("h2", fontName=font_name, fontSize=11, leading=16,
                             textColor=colors.HexColor(f"#1F4E79"), spaceBefore=10)

    story = []

    # 표지
    story.append(Paragraph("AP 자동검증 결과보고서", title_s))
    story.append(Spacer(1, 4*mm))

    total_pass = sum(r.pass_count for r in records)
    total_fail = sum(r.fail_count for r in records)
    total_all  = total_pass + total_fail
    rate       = f"{total_pass/total_all*100:.1f}%" if total_all else "—"
    judged     = total_fail == 0

    info_data = [
        ["항목","내용"],
        ["시나리오", scenario.name],
        ["플러그인", scenario.plugin],
        ["장비명",   device_info.get("device_name","—")],
        ["펌웨어",   device_info.get("firmware","—")],
        ["담당자",   device_info.get("tester","—")],
        ["일시",     datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["총 반복",  str(len(records))],
        ["합격률",   rate],
        ["최종 판정","PASS" if judged else "FAIL"],
    ]
    tbl = Table(info_data, colWidths=[40*mm, 100*mm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(-1,0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR",  (0,0),(-1,0), colors.white),
        ("FONTNAME",   (0,0),(-1,-1), font_name),
        ("FONTSIZE",   (0,0),(-1,-1), 9),
        ("GRID",       (0,0),(-1,-1), 0.5, colors.HexColor("#BFBFBF")),
        ("ROWBACKGROUNDS",(0,1),(-1,-2),[colors.white, colors.HexColor("#F2F2F2")]),
        ("BACKGROUND", (0,-1),(-1,-1),
         colors.HexColor("#E2EFDA") if judged else colors.HexColor("#FCE4D6")),
        ("TEXTCOLOR",  (0,-1),(-1,-1),
         colors.HexColor("#375623") if judged else colors.HexColor("#9C0006")),
        ("FONTNAME",   (0,-1),(-1,-1), font_name),
        ("ALIGN",      (0,0),(-1,-1), "CENTER"),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 6*mm))

    # 상세 결과
    story.append(Paragraph("Step별 상세 결과", h2_s))
    detail_header = ["회차","Step","유형","항목명","기댓값","실측값","ms","판정"]
    detail_data = [detail_header]
    for rec in records:
        for sr in rec.step_results:
            detail_data.append([
                str(rec.iteration),
                str(sr.step_index+1),
                sr.step_type,
                sr.step_name[:20],
                str(sr.expected or "—")[:12],
                str(sr.actual   or "—")[:12],
                f"{sr.elapsed_ms:.0f}",
                "PASS" if sr.passed else "FAIL",
            ])

    col_w = [12*mm,12*mm,14*mm,45*mm,25*mm,25*mm,12*mm,12*mm]
    dtbl  = Table(detail_data, colWidths=col_w, repeatRows=1)
    style = [
        ("BACKGROUND", (0,0),(-1,0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR",  (0,0),(-1,0), colors.white),
        ("FONTNAME",   (0,0),(-1,-1), font_name),
        ("FONTSIZE",   (0,0),(-1,-1), 8),
        ("GRID",       (0,0),(-1,-1), 0.5, colors.HexColor("#BFBFBF")),
        ("ALIGN",      (0,0),(-1,-1), "CENTER"),
    ]
    for i, rec in enumerate(records):
        for j, sr in enumerate(rec.step_results):
            row_i = 1 + sum(len(r.step_results) for r in records[:i]) + j
            bg = colors.HexColor("#E2EFDA") if sr.passed else colors.HexColor("#FCE4D6")
            style.append(("BACKGROUND", (7,row_i),(7,row_i), bg))
    dtbl.setStyle(TableStyle(style))
    story.append(dtbl)

    # FAIL 목록
    fails = [(rec, sr) for rec in records for sr in rec.step_results if not sr.passed]
    if fails:
        story.append(Spacer(1,6*mm))
        story.append(Paragraph("FAIL 항목 목록", h2_s))
        fail_data = [["회차","Step","항목명","메시지"]]
        for rec, sr in fails:
            fail_data.append([str(rec.iteration), str(sr.step_index+1),
                              sr.step_name[:18], sr.message[:30]])
        ftbl = Table(fail_data, colWidths=[12*mm,12*mm,45*mm,88*mm])
        ftbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#9C0006")),
            ("TEXTCOLOR", (0,0),(-1,0),colors.white),
            ("FONTNAME",  (0,0),(-1,-1),font_name),
            ("FONTSIZE",  (0,0),(-1,-1),8),
            ("GRID",      (0,0),(-1,-1),0.5,colors.HexColor("#BFBFBF")),
            ("BACKGROUND",(0,1),(-1,-1),colors.HexColor("#FCE4D6")),
        ]))
        story.append(ftbl)

    doc.build(story)
    return path
